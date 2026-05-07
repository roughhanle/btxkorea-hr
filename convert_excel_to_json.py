"""BTXKOREA 스킬관리앱 — 엑셀(v14) → JSON 변환 스크립트.

산출:
  data/skills.json        — 78개 스킬 정의
  data/questions.json     — 702개 문항 (선택지 + EPDN 매핑)
  data/matrix.json        — 직책·직무 매트릭스 룰
  data/scoring_rules.json — 점수 산출 룰 (시트 17 기반 상수)
  data/prescriptions.json — 20개 육성 처방
  data/guide.json         — 시트 1(개요)·시트 2(작성자 안내) 본문

사용:  python convert_excel_to_json.py
"""

from __future__ import annotations
import json
from pathlib import Path
import openpyxl

ROOT = Path(__file__).parent
XLSX = ROOT.parent / "비티엑스코리아_스킬관리앱DB_v14.xlsx"
OUT_DIR = ROOT / "data"
OUT_DIR.mkdir(exist_ok=True)

JOB_CODE = {
    "사외물류": "WO",
    "사내물류": "IL",
    "조립·포장": "AP",
    "운송납품": "TR",
    "통합": "IN",
}

# 스킬 시트와 매핑
SKILL_SHEETS = [
    ("4. 공통 스킬 라이브러리", "공통"),
    ("5. 직무 스킬_사외물류", "사외물류"),
    ("6. 직무 스킬_사내물류", "사내물류"),
    ("7. 직무 스킬_조립·포장", "조립·포장"),
    ("8. 직무 스킬_운송납품", "운송납품"),
    ("9. 통합 트랙 스킬", "통합"),
]

# 문항 시트와 매핑
QUESTION_SHEETS = [
    ("11. 문항_공통", "공통", False),     # 16컬럼 (직무 없음)
    ("12. 문항_사외물류", "사외물류", True),
    ("13. 문항_사내물류", "사내물류", True),
    ("14. 문항_조립·포장", "조립·포장", True),
    ("15. 문항_운송납품", "운송납품", True),
    ("16. 문항_통합트랙", "통합", True),
]


def cell_str(v):
    """엑셀 셀 값을 깔끔한 문자열로 변환."""
    if v is None:
        return ""
    return str(v).strip()


# -------------------------------------------------------------------
# 1. skills.json
# -------------------------------------------------------------------
def build_skills(wb):
    skills = []
    for sheet_name, job in SKILL_SHEETS:
        ws = wb[sheet_name]
        is_common = job == "공통"
        # 헤더는 R4. 데이터는 R5부터
        for r in range(5, ws.max_row + 1):
            skill_id = cell_str(ws.cell(r, 1).value)
            if not skill_id:
                continue
            if is_common:
                # 8컬럼: ID, 카테고리, 구분, 스킬명, 정의, 기술, 지식, 관리툴
                row = {
                    "skill_id": skill_id,
                    "category": cell_str(ws.cell(r, 2).value),
                    "kind": cell_str(ws.cell(r, 3).value),
                    "job": "공통",
                    "skill_name": cell_str(ws.cell(r, 4).value),
                    "definition": cell_str(ws.cell(r, 5).value),
                    "skills": cell_str(ws.cell(r, 6).value),
                    "knowledge": cell_str(ws.cell(r, 7).value),
                    "tools": cell_str(ws.cell(r, 8).value),
                }
            else:
                # 9컬럼: ID, 카테고리, 구분, 직무, 스킬명, 정의, 기술, 지식, 관리툴
                row = {
                    "skill_id": skill_id,
                    "category": cell_str(ws.cell(r, 2).value),
                    "kind": cell_str(ws.cell(r, 3).value),
                    "job": cell_str(ws.cell(r, 4).value),
                    "skill_name": cell_str(ws.cell(r, 5).value),
                    "definition": cell_str(ws.cell(r, 6).value),
                    "skills": cell_str(ws.cell(r, 7).value),
                    "knowledge": cell_str(ws.cell(r, 8).value),
                    "tools": cell_str(ws.cell(r, 9).value),
                }
            skills.append(row)
    return skills


# -------------------------------------------------------------------
# 2. questions.json
# -------------------------------------------------------------------
def build_questions(wb):
    questions = []
    for sheet_name, job, has_job_col in QUESTION_SHEETS:
        ws = wb[sheet_name]
        for r in range(5, ws.max_row + 1):
            skill_id = cell_str(ws.cell(r, 1).value)
            if not skill_id:
                continue

            if has_job_col:
                # 17컬럼: ID, 카테고리, 스킬명, 직무, 직책, 문항#, 역할, 문항ID, 문항텍스트, 선택지×4 + 매핑×4
                category = cell_str(ws.cell(r, 2).value)
                skill_name = cell_str(ws.cell(r, 3).value)
                job_in = cell_str(ws.cell(r, 4).value)
                role = cell_str(ws.cell(r, 5).value)
                qnum = ws.cell(r, 6).value
                q_role = cell_str(ws.cell(r, 7).value)
                qid = cell_str(ws.cell(r, 8).value)
                qtext = cell_str(ws.cell(r, 9).value)
                choice_cols = [(10, 11), (12, 13), (14, 15), (16, 17)]
            else:
                # 16컬럼 (공통, 직무 컬럼 없음): ID, 카테고리, 스킬명, 직책, 문항#, 문항역할, 문항ID, 문항텍스트, 선택지×4 + 매핑×4
                category = cell_str(ws.cell(r, 2).value)
                skill_name = cell_str(ws.cell(r, 3).value)
                job_in = "공통"
                role = cell_str(ws.cell(r, 4).value)
                qnum = ws.cell(r, 5).value
                q_role = cell_str(ws.cell(r, 6).value)
                qid = cell_str(ws.cell(r, 7).value)
                qtext = cell_str(ws.cell(r, 8).value)
                choice_cols = [(9, 10), (11, 12), (13, 14), (15, 16)]

            if not qid:
                continue

            labels = ["①", "②", "③", "④"]
            choices = []
            for i, (text_col, epdn_col) in enumerate(choice_cols):
                choices.append({
                    "label": labels[i],
                    "text": cell_str(ws.cell(r, text_col).value),
                    "epdn": cell_str(ws.cell(r, epdn_col).value).upper(),
                })

            questions.append({
                "question_id": qid,
                "skill_id": skill_id,
                "category": category,
                "skill_name": skill_name,
                "job": job_in,
                "role": role,             # 팀장/사업부장/본부장
                "question_number": int(qnum) if qnum is not None else None,
                "question_role": q_role,  # 행동/산출물/수준
                "question_text": qtext,
                "choices": choices,
            })
    return questions


# -------------------------------------------------------------------
# 3. matrix.json — 시트 10 룰을 코드가 사용하기 좋은 형태로 정리
# -------------------------------------------------------------------
def build_matrix():
    """시트 10의 매트릭스를 앱 분기 로직용 구조로 정규화."""
    return {
        "role_job_limits": {
            "팀장":   {"min": 1, "max": 2, "auto_integrated_threshold": 3},
            "사업부장": {"min": 1, "max": 3, "auto_integrated_threshold": 4},
            "본부장":  {"auto_integrated": True},
        },
        "role_codes": {"팀장": "T", "사업부장": "D", "본부장": "H"},
        "job_codes": JOB_CODE,
        "categories": ["Ⅰ. 공통", "Ⅱ. 설계", "Ⅲ. 운영", "Ⅳ. 평가", "Ⅴ. 개선"],
        # 케이스별 스킬 선택 룰 (실제 스킬 ID는 코드에서 skills.json + 이 룰로 조립)
        "skill_selection": {
            "case_single_job": {
                "description": "팀장·사업부장 1개 직무 선택",
                "Ⅰ. 공통": {"common_skills": ["C-01", "C-02", "C-03", "C-04", "C-05", "C-06"], "job_skills": []},
                "Ⅱ. 설계": {"common_skills": ["S-01", "S-02", "S-03"], "job_skill_pattern": "S-{job_code}-{n}", "n_range": [1, 2, 3]},
                "Ⅲ. 운영": {"common_skills": ["O-01", "O-02", "O-03"], "job_skill_pattern": "O-{job_code}-{n}", "n_range": [1, 2, 3]},
                "Ⅳ. 평가": {"common_skills": ["E-01", "E-02", "E-03"], "job_skill_pattern": "E-{job_code}-{n}", "n_range": [1, 2, 3]},
                "Ⅴ. 개선": {"common_skills": ["I-01", "I-02", "I-03"], "job_skill_pattern": "I-{job_code}-{n}", "n_range": [1, 2, 3]},
            },
            "case_two_jobs": {
                "description": "팀장·사업부장 2개 직무 선택 — 공통 직무카테고리 스킬 빠짐",
                "Ⅰ. 공통": {"common_skills": ["C-01", "C-02", "C-03", "C-04", "C-05", "C-06"], "job_skills": []},
                "Ⅱ. 설계": {"common_skills": [], "job_skill_pattern": "S-{job_code}-{n}", "n_range": [1, 2, 3]},
                "Ⅲ. 운영": {"common_skills": [], "job_skill_pattern": "O-{job_code}-{n}", "n_range": [1, 2, 3]},
                "Ⅳ. 평가": {"common_skills": [], "job_skill_pattern": "E-{job_code}-{n}", "n_range": [1, 2, 3]},
                "Ⅴ. 개선": {"common_skills": [], "job_skill_pattern": "I-{job_code}-{n}", "n_range": [1, 2, 3]},
            },
            "case_integrated": {
                "description": "본부장 / 3개 이상 직무 — 통합 트랙 자동 적용",
                "Ⅰ. 공통": {"common_skills": ["C-01", "C-02", "C-03", "C-04", "C-05", "C-06"], "integrated_skills": []},
                "Ⅱ. 설계": {"common_skills": ["S-01", "S-02", "S-03"], "integrated_skills": ["S-IN-01", "S-IN-02", "S-IN-03"]},
                "Ⅲ. 운영": {"common_skills": ["O-01", "O-02", "O-03"], "integrated_skills": ["O-IN-01", "O-IN-02", "O-IN-03"]},
                "Ⅳ. 평가": {"common_skills": ["E-01", "E-02", "E-03"], "integrated_skills": ["E-IN-01", "E-IN-02", "E-IN-03"]},
                "Ⅴ. 개선": {"common_skills": ["I-01", "I-02", "I-03"], "integrated_skills": ["I-IN-01", "I-IN-02", "I-IN-03"]},
            },
        },
        "role_target_epdn": {
            "본부장":   {"Ⅰ. 공통": "E", "Ⅱ. 설계": "P", "Ⅲ. 운영": "P", "Ⅳ. 평가": "P", "Ⅴ. 개선": "E"},
            "사업부장": {"Ⅰ. 공통": "P", "Ⅱ. 설계": "P", "Ⅲ. 운영": "P", "Ⅳ. 평가": "P", "Ⅴ. 개선": "P"},
            "팀장":    {"Ⅰ. 공통": "P", "Ⅱ. 설계": "D", "Ⅲ. 운영": "P", "Ⅳ. 평가": "D", "Ⅴ. 개선": "D"},
        },
        "divisions": ["군산", "울산", "인천", "중부", "창원", "본사"],
    }


# -------------------------------------------------------------------
# 4. scoring_rules.json — 시트 17 룰
# -------------------------------------------------------------------
def build_scoring_rules():
    return {
        "epdn_score": {"E": 4, "P": 3, "D": 2, "N": 1},
        "skill_score": {
            "formula": "round((q1 + q2 + q3) / 3, 2)",
            "min": 1.00, "max": 4.00,
        },
        "category_score": {
            "formula": "sum(6 skill scores)",
            "min": 6, "expected": 18, "max": 24,
        },
        "total_score": {
            "formula": "sum(5 category scores)",
            "min": 30, "expected": 90, "max": 120,
        },
        "achievement_rate": {
            "formula": "round((total / 90) * 100, 1)",
            "expected_baseline": 90,
        },
        "overall_grade_thresholds": [
            {"grade": "우수",      "min_rate": 120, "color": "#C8E6C9"},
            {"grade": "안정",      "min_rate": 100, "color": "#BBDEFB"},
            {"grade": "보완 필요",  "min_rate": 80,  "color": "#FFE082"},
            {"grade": "집중 육성",  "min_rate": 0,   "color": "#FFCDD2"},
        ],
        "category_grade_thresholds": [
            # 카테고리 등급은 카테고리 달성률(점수/18*100) 기준이지만 임계값은 동일
            {"grade": "우수",      "min_rate": 120, "min_score": 21.6},
            {"grade": "안정",      "min_rate": 100, "min_score": 18.0},
            {"grade": "보완 필요",  "min_rate": 80,  "min_score": 14.4},
            {"grade": "집중 육성",  "min_rate": 0,   "min_score": 0.0},
        ],
        "prescription_id_map": {
            "Ⅰ. 공통": "COM",
            "Ⅱ. 설계": "DES",
            "Ⅲ. 운영": "OPS",
            "Ⅳ. 평가": "EVL",
            "Ⅴ. 개선": "IMP",
        },
        "grade_code_map": {
            "우수": "E",
            "안정": "S",
            "보완 필요": "N",
            "집중 육성": "F",
        },
        "grade_messages": {
            "우수":      "직책 기대를 상회 — 리더십 확장·후임 멘토링 역할 부여 권장",
            "안정":      "직책 기대 충족 — 약점 카테고리 집중 보완·E 수준 도전 과제",
            "보완 필요":  "직책 기대 미달 — OJT 강화·코칭 정기화·표준 매뉴얼 학습",
            "집중 육성":  "기본기 부족 — 기본 직무 교육·1:1 멘토링·점진적 책임 부여",
        },
    }


# -------------------------------------------------------------------
# 5. prescriptions.json — 시트 18
# -------------------------------------------------------------------
# -------------------------------------------------------------------
# 5b. guide.json — 시트 1(개요)·시트 2(작성자 안내) 본문
# -------------------------------------------------------------------
def _extract_guide_sheet(ws):
    """행 단위 셀 배열을 순서대로 수집. 빈 행은 새 줄(빈 cells) 마커로 보존."""
    rows = []
    for r in range(1, ws.max_row + 1):
        cells = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is not None:
                cells.append(str(v).strip())
        rows.append({"row": r, "cells": cells})
    # 끝에 붙은 빈 행 정리
    while rows and not rows[-1]["cells"]:
        rows.pop()
    return rows


def build_guide(wb):
    return {
        "overview": _extract_guide_sheet(wb["1. 개요"]),
        "respondent_guide": _extract_guide_sheet(wb["2. 작성자 안내"]),
    }


def build_prescriptions(wb):
    ws = wb["18. 육성방안 라이브러리"]
    prescriptions = []
    # 헤더 R5, 데이터 R6~R25
    for r in range(6, ws.max_row + 1):
        pid = cell_str(ws.cell(r, 1).value)
        if not pid:
            continue
        prescriptions.append({
            "prescription_id": pid,
            "category": cell_str(ws.cell(r, 2).value),
            "grade": cell_str(ws.cell(r, 3).value),
            "score_range": cell_str(ws.cell(r, 4).value),
            "diagnosis": cell_str(ws.cell(r, 5).value),
            "ojt_70": cell_str(ws.cell(r, 6).value),
            "coaching_20": cell_str(ws.cell(r, 7).value),
            "education_10": cell_str(ws.cell(r, 8).value),
            "period": cell_str(ws.cell(r, 9).value),
            "note": cell_str(ws.cell(r, 10).value),
        })
    return prescriptions


# -------------------------------------------------------------------
# main
# -------------------------------------------------------------------
def write_json(name, data):
    path = OUT_DIR / name
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def main():
    print(f"Loading {XLSX.name}...")
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    skills = build_skills(wb)
    questions = build_questions(wb)
    matrix = build_matrix()
    scoring = build_scoring_rules()
    prescriptions = build_prescriptions(wb)
    guide = build_guide(wb)

    write_json("skills.json", skills)
    write_json("questions.json", questions)
    write_json("matrix.json", matrix)
    write_json("scoring_rules.json", scoring)
    write_json("prescriptions.json", prescriptions)
    write_json("guide.json", guide)

    # 검증 출력
    print(f"\n=== Conversion summary ===")
    print(f"  skills.json        : {len(skills):4d} skills")
    print(f"  questions.json     : {len(questions):4d} questions")
    print(f"  matrix.json        : {len(matrix)} top-level keys")
    print(f"  scoring_rules.json : {len(scoring)} top-level keys")
    print(f"  prescriptions.json : {len(prescriptions):4d} prescriptions")

    # 검증: 기대 건수
    print(f"\n=== Expected vs actual ===")
    print(f"  skills      expected 78  → {len(skills)}  {'OK' if len(skills)==78 else 'MISMATCH'}")
    print(f"  questions   expected 702 → {len(questions)}  {'OK' if len(questions)==702 else 'MISMATCH'}")
    print(f"  prescriptions expected 20 → {len(prescriptions)}  {'OK' if len(prescriptions)==20 else 'MISMATCH'}")

    # 카테고리 분포 점검
    from collections import Counter
    cat_counts = Counter(s["category"] for s in skills)
    print(f"\n  skills by category: {dict(cat_counts)}")
    job_counts = Counter(s["job"] for s in skills)
    print(f"  skills by job: {dict(job_counts)}")

    # 문항 EPDN 검증 — 모든 문항이 4개 선택지 모두 EPDN 매핑되었는지
    bad_q = []
    for q in questions:
        epdns = [c["epdn"] for c in q["choices"]]
        if not all(e in {"E", "P", "D", "N"} for e in epdns):
            bad_q.append((q["question_id"], epdns))
    if bad_q:
        print(f"  WARNING: {len(bad_q)} questions have invalid EPDN mappings")
        for qid, epdns in bad_q[:5]:
            print(f"    {qid}: {epdns}")
    else:
        print(f"  EPDN mapping: all {len(questions)} questions OK")


if __name__ == "__main__":
    main()
